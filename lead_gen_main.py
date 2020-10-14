import requests
import openpyxl
import lxml
from bs4 import BeautifulSoup
from scrape import website_grab, smoking_grab, social_grab

print("\nMake sure your path.txt is properly addressed! Otherwise the application will crash.\n")
path_txt = open('path.txt', 'r')
path = path_txt.read()
path_txt.close()
data_dict = {}
n = 20
offset_var = 0
set_loc = input("Enter location to generate leads: ")
set_category = input("\n\nEnter category term exactly as listed in cat_list.pdf i.e: danceclubs"
                     "\nIf you want multiple categories, type them separated by commas WITHOUT a space,"
                     " i.e: danceclubs,bars\n\nEnter Category: ")

while n >= 1:

    API_KEY = 'Uok93zaKIflvNPgElgAykwI5ptxiVPC_Zwt0mDLb7vtrf0tf2eF9F5qZkIIx5LqjrlwwTHR71TdzU3mgBawIq1em37fH' \
              'LnYHpKOrcI0miOD-iX-tdIpPAMQGFXt6X3Yx'
    ENDPOINT = 'https://api.yelp.com/v3/businesses/search'
    HEADERS = {'Authorization': 'bearer %s' % API_KEY}

    PARAMETERS = {'categories': '%s' % set_category,
                  'limit': 25,
                  'radius': 40000,
                  'offset': offset_var,
                  'location': '%s' % set_loc,
                  }
    response = requests.get(url=ENDPOINT,
                            params=PARAMETERS,
                            headers=HEADERS)

    business_data = response.json()
    offset_var = offset_var + 25
    print(str(100-((n * 100) / 20)) + "% finished")
    n = n - 1
    for i in business_data['businesses']:
        name = i["name"]
        print(name)
        review_count = i["review_count"]
        rating = i["rating"]
        phone = i["display_phone"]
        id = i["id"]
        url = i["url"]
        headers = {
            'User-Agent': 'Chrome/85.0.4183.121',
        }
        source = requests.get('%s' % url, headers=headers).text
        soup = BeautifulSoup(source, 'lxml')
        website = soup.get
        website = website_grab(website)
        smoking = soup.get
        smoking = smoking_grab(smoking)
        try:
            if website != "No business website detected":
                http_web = 'http://%s' % website + "/index/home"
                source = requests.get('%s' % http_web, headers=headers).text
                soup = BeautifulSoup(source, 'lxml')
                social = soup.get
                social = social_grab(social)
                if social[0] and social[1] == "N/A":
                    http_web = 'http://%s' % website
                    source = requests.get('%s' % http_web, headers=headers).text
                    soup = BeautifulSoup(source, 'lxml')
                    social = soup.get
                    social = social_grab(social)
                else:
                    pass
            else:
                social = ["N/A", "N/A"]
        except:
            social = ["Web error", "Web error"]
        print("Parsing...")
        try:
            price = i["price"]
        except:
            price = "N/A"
        try:
            address = i["location"]["address1"] + ", " + set_loc
        except:
            address = "N/a"
        data_dict[name] = [review_count, rating, phone, price, address, smoking[0], smoking[1], website, social[0],
                           social[1], url]
print("100% finished!")
print("Exporting to excel...")
workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1, value="Name of Business")
sheet.cell(row=1, column=2, value="# of Reviews")
sheet.cell(row=1, column=3, value="Avg Review")
sheet.cell(row=1, column=4, value="phone #")
sheet.cell(row=1, column=5, value="Price level ")
sheet.cell(row=1, column=6, value="Address in " + set_loc)
sheet.cell(row=1, column=7, value="Smoking allowed inside?")
sheet.cell(row=1, column=8, value="Smoking allowed outside?")
sheet.cell(row=1, column=9, value="Business Website")
sheet.cell(row=1, column=10, value="Facebook")
sheet.cell(row=1, column=11, value="Instagram")
sheet.cell(row=1, column=12, value="Yelp url")

row = 2
for key, values in data_dict.items():
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 2

workbook.save('%s/Leads_generated' % path + '_%s_' % set_loc + '_%s_.xlsx' % set_category)
